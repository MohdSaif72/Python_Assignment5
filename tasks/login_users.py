"""
Description:
This script defines a UserLogin class to manage user login operations using Selenium. 
It loads user credentials from an Excel file, performs login attempts on a web page, captures
error messages, and stores the results in another Excel sheet. The class utilizes explicit waits
for element interactions and URL changes during the login process.

Classes:
- UserLogin: Manages user login operations and result storage.

Functions/Methods:
- __init__: Initializes the LoginManager class with URL, filename, and browser manager.
- login_users: Performs user login attempts and stores results in an Excel file.
"""
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import openpyxl
from login import Login

class UserLogin:
    """
        Initialize the UserLogin class with the specified URL and filename for logging in users and recording login messages.
        
        Parameters:
        url (str): The URL of the website for user login.
        filename (str): The name of the Excel file containing user credentials and login messages.
        browser_manager: An instance of BrowserManager to control the web browser.
        """
    def __init__(self, url, filename,browser_manager):
        self.url = url
        self.filename = filename
        # self.driver = webdriver.Chrome()
        self.browser_manager=browser_manager
             
    def login_users(self):
        """
        This method performs login with multiple users and records login messages.
        """
        try:
            workbook = openpyxl.load_workbook(self.filename)
            login_sheet = workbook.create_sheet("Login")
            login_sheet.append(["User ID", "Login Message"])
            users = workbook["User credentials"]
                  
            for row in users.iter_rows(min_row=2, values_only=True):
                user_id, username, password = row
                self.browser_manager.setup_browser(self.url)
                login_manager = Login(self.browser_manager,username,password)
                login_manager.login()
               
                initial_url = self.browser_manager.driver.current_url
                if self.browser_manager.driver.current_url == initial_url:
                    pass
                else:
                    self.browser_manager.driver.execute_script("window.history.go(-1)")              
                error_message = ""
                try:
                    error_message = self.browser_manager.driver.find_element(By.XPATH, "//h3").text
                except NoSuchElementException as e:
                    error_message = f"Element not found: {e}"
                                     
                login_sheet.append([user_id, error_message])
            workbook.save(self.filename)

        finally:
            self.browser_manager.driver.quit()



