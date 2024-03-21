from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def new_sheet(filepath, sheetname, columnheaders):
    """
    Create a new sheet in an existing Excel file with the specified column headers.

    Parameters:
    - filepath (str): Path to the existing Excel file.
    - sheetname (str): Name of the new sheet to be created.
    - columnheaders (list): List of column headers to be added to the new sheet.
    """
    try:
        wb = load_workbook(filename=filepath)
        if sheetname in wb.sheetnames:
            return
        
        new_sheet = wb.create_sheet(title=sheetname)
        
        for col_idx, header in enumerate(columnheaders, start=1):
            column_letter = get_column_letter(col_idx)
            new_sheet[f"{column_letter}1"] = header
        
        wb.save(filename=filepath)
    except Exception as e:
        print(f"An error occurred: {e}")

filepath = "saucedemo.xlsx"
sheetname = "Order Details"
columnheaders = ["User ID", "Product Name", "Quantity", "Price"]
new_sheet(filepath, sheetname, columnheaders)
