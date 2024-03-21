"""
The selenium package is used to automate web browser interaction.
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
import time

class UserLogin:
    """
        Initialize the UserLogin class with the specified URL and filename for logging in users and recording login messages.
        
        Parameters:
        url (str): The URL of the website for user login.
        filename (str): The name of the Excel file containing user credentials and login messages.
        """
    def __init__(self, url, filename):
        self.url = url
        self.filename = filename
        self.driver = webdriver.Chrome()

    def login_users(self):
        """
        This method performs login with multiple users and records login messages.
        """
        try:
            workbook = openpyxl.load_workbook(self.filename)
            login_sheet = workbook.create_sheet("Login")
            login_sheet.append(["User ID", "Login Message"])
            users = workbook["User credentials"]
            
            for row in users.iter_rows(min_row=2, values_only=True):
                user_id, username, password = row
                self.driver.get(self.url)
                username_field = self.driver.find_element(By.ID, "user-name")
                password_field = self.driver.find_element(By.ID, "password")
                username_field.send_keys(username)
                password_field.send_keys(password)
                password_field.send_keys(Keys.RETURN)
                time.sleep(5) 
                error_message = ""
                try:
                    error_message = self.driver.find_element(By.XPATH, "//h3").text
                except:
                    pass
                login_sheet.append([user_id, error_message])
            workbook.save(self.filename)

        finally:
            self.driver.quit()



