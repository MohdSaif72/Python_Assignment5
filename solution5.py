import openpyxl

def add_column(filename, sheetname, column_header):
    """
     Add an empty column at the end of a specific sheet in an Excel file.

    Parameters:
    - file_name (str): The name of the Excel file.
    - sheet_name (str): The name of the specific sheet where the empty column will be added.
    - new_column_header (str): The header for the new empty column.
    """
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    sheet.cell(row=1, column=sheet.max_column + 1, value=column_header)
    workbook.save(filename)

# Call the function with the appropriate arguments
add_column("saucedemo.xlsx", "Order Details", "Order Status")

