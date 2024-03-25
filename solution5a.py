"""Openpyxl is a Python library for reading and writing Excel files. 
"""
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time 

class Order:
    """
    A class to place orders using data from an Excel file.

    Attributes:
    - filename (str): The name of the Excel file containing order details.
    - website_url (str): The URL of the website to place orders on.
    """

    def __init__(self, filename, url):
        self.filename = filename
        self.url = url

    def orders(self):
        """
        Places orders based on data from the specified Excel file.
        """
        try:
            workbook = openpyxl.load_workbook(self.filename)
            print("Sheet names:", workbook.sheetnames)
            order_sheet = workbook["Order Details"]
            order_sheet["E1"] = "Order Status"  
            rows = order_sheet.iter_rows(min_row=2, values_only=True)
            if rows is None:
                print("No rows found in the worksheet.")
                return

            # Iterate over each row
            for row_index, row in enumerate(rows, start=1):
                username, product_name, quantity, *_ = row

                driver = webdriver.Chrome()

                try:
                    driver.get(self.url)  
                    print("Opened website successfully")
                    
                    # Locate elements for login
                    username_field = driver.find_element(By.ID, "user-name")
                    password_field = driver.find_element(By.ID, "password")
                    login_button = driver.find_element(By.CLASS_NAME, "btn_action")

                    username_field.send_keys(username)
                    password_field.send_keys("secret_sauce")
                    login_button.click()
                    print("Logged in successfully")
                    
                    driver.implicitly_wait(5)
                    time.sleep(3)

                    # Add product to cart
                    product = driver.find_element(By.XPATH, f"//*[text()='{product_name}']/../../..")
                    add_to_cart_button = product.find_element(By.CLASS_NAME, "btn_inventory")
                    add_to_cart_button.click()
                    print("Product added to cart successfully")
                    time.sleep(3)

                    # Verify if product added to cart successfully
                    try:
                        cart_icon = driver.find_element(By.XPATH, "//a[@class='shopping_cart_link fa-layers fa-fw']")
                        cart_icon.click()
                    except Exception as e:
                        print("Error occurred while retrieving cart count:", e)
                        continue  # Skip to the next iteration if there's an error

                    # Retrieve the product name from the cart
                    try:
                        cart_product_name = driver.find_element(By.XPATH, "//*[@id='item_4_title_link']/div").text
                        print("Product name in cart:", cart_product_name)
                    except Exception as e:
                        print("Error occurred while retrieving product name from cart:", e)
                        continue  # Skip to the next iteration if there's an error

                    # Compare the product name in the cart with the expected product name
                    if cart_product_name == product_name:
                        order_sheet.cell(row=order_sheet.min_row + row_index - 1, column=5, value="Success")  # Update status in the last column for the current row
                        print("Order status updated to Success")
                    else:
                        order_sheet.cell(row=order_sheet.min_row + row_index - 1, column=5, value="Failure")  # Update status in the last column for the current row
                        print("Order status updated to Failure")

                finally:
                    workbook.save(self.filename)
                    print("Excel file saved successfully")
                    driver.quit()  

        except Exception as e:
            print("An error occurred:", e)

# Call place_orders function with provided filename and website URL
filename = "saucedemo.xlsx"
website_url = "https://www.saucedemo.com/v1/"
order_placer = Order(filename, website_url)
order_placer.orders()
